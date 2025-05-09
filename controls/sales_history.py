import mariadb
import pandas as pd
import os
from PyQt6 import uic
from decimal import Decimal
from db.config import db_config
from PyQt6.QtWidgets import (
    QMainWindow, QLineEdit, QCalendarWidget, QTableWidget, QTableWidgetItem,
    QPushButton, QMessageBox, QProgressDialog, QLabel
)
from PyQt6.QtCore import QDate, Qt, QThread, pyqtSignal
#excel
import openpyxl
from openpyxl.styles import Alignment
#pdf
from fpdf import FPDF
class SalesHistoryWindow(QMainWindow):
    def __init__(self, user_id, db_config, dashboard_window):
        super().__init__()
        uic.loadUi("ui/sales_history.ui", self)

        self.user_id = user_id
        self.db_config = db_config
        self.dashboard_window = dashboard_window

        #connnection  to the ui elements
        self.calendar = self.findChild(QCalendarWidget, "calendarWidget")
        self.sales_table = self.findChild(QTableWidget, "salesTable")
        self.export_excel_button = self.findChild(QPushButton, "exportExcelButton")
        self.export_pdf_button = self.findChild(QPushButton, "exportPdfButton")
        self.back_button = self.findChild(QPushButton, "backButton")
        self.total_purchase_label = self.findChild(QLabel, "totalPurchaseLabel")
        self.total_sales_label = self.findChild(QLabel, "totalSalesLabel")
        self.total_income_label = self.findChild(QLabel, "totalIncomeLabel")
        self.date_label = self.findChild(QLabel, "dateLabel")

        self.search_history = self.findChild(QLineEdit, "searchHistory")
        self.search_history.setClearButtonEnabled(True)
        self.search_history.textChanged.connect(self.search_product)

        self.sales_table.setColumnCount(5)
        self.sales_table.setHorizontalHeaderLabels([
            "Order ID", "Product Name", "Quantity", "Total Retail Sales", "Sales Date"
        ])

        current_date = QDate.currentDate().toString("yyyy-MM-dd")
        self.date_label.setText(f"Date now: {current_date}")

        self.calendar.selectionChanged.connect(self.load_sales)
        self.export_excel_button.clicked.connect(self.export_to_excel)
        self.export_pdf_button.clicked.connect(self.export_to_pdf)
        self.back_button.clicked.connect(self.go_back)

        self.calendar.setSelectedDate(QDate.currentDate())
        self.load_sales_for_today()

    def go_back(self):
        self.dashboard_window.show()
        self.close()

    def load_sales_for_today(self):
        selected_date = QDate.currentDate().toString("yyyy-MM-dd")
        self.load_sales(selected_date)

    def load_sales(self, selected_date=None):
        if selected_date is None:
            selected_date = self.calendar.selectedDate().toString("yyyy-MM-dd")

        self.loading_dialog = QProgressDialog("Loading sales data...", "Cancel", 0, 0, self)
        self.loading_dialog.setWindowModality(Qt.WindowModality.WindowModal)
        self.loading_dialog.setCancelButton(None)
        self.loading_dialog.setValue(0)
        self.loading_dialog.show()

        self.thread = SalesLoaderThread(self.user_id, selected_date, self.db_config)
        self.thread.finished.connect(self.on_sales_data_loaded)
        self.thread.start()

    def on_sales_data_loaded(self, sales_data):
        self.loading_dialog.close()
        self.sales_table.setRowCount(0)

        if not sales_data:
            QMessageBox.warning(self, "No Data", "No sales data found.")
            self.total_purchase_label.setText("Total Purchase: 0.00")
            self.total_sales_label.setText("Total Sales: 0.00")
            self.total_income_label.setText("Total Income: 0.00")
            return

        self.order_sales = {}
        total_purchase = Decimal("0.00")
        total_sales = Decimal("0.00")

        for sale in sales_data:
            order_id, product_name, quantity, total_price, order_datetime, purchase_price = sale
            total_price = Decimal(total_price)
            purchase_price = Decimal(purchase_price)

            total_sales += total_price
            total_purchase += purchase_price * Decimal(quantity)

            if order_id not in self.order_sales:
                self.order_sales[order_id] = {
                    "products": [],
                    "total_sales": Decimal("0.00"),
                    "sales_date": order_datetime.date()
                }

            self.order_sales[order_id]["products"].append((product_name, quantity))
            self.order_sales[order_id]["total_sales"] += total_price

        tallest_row_height = 0

        for order_id, details in self.order_sales.items():
            products = ", ".join([p[0] for p in details["products"]])
            quantities = ", ".join([str(p[1]) for p in details["products"]])
            total = details["total_sales"]
            sales_date = details["sales_date"]

            row_position = self.sales_table.rowCount()
            self.sales_table.insertRow(row_position)

            product_item = QTableWidgetItem(products)
            product_item.setTextAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft)
            self.sales_table.setItem(row_position, 1, product_item)

            self.sales_table.setItem(row_position, 0, QTableWidgetItem(str(order_id)))
            self.sales_table.setItem(row_position, 2, QTableWidgetItem(quantities))
            self.sales_table.setItem(row_position, 3, QTableWidgetItem(f"{total:.2f}"))
            self.sales_table.setItem(row_position, 4, QTableWidgetItem(str(sales_date)))

            self.sales_table.resizeRowsToContents()
            row_height = self.sales_table.rowHeight(row_position)
            if row_height > tallest_row_height:
                tallest_row_height = row_height

        for row in range(self.sales_table.rowCount()):
            self.sales_table.setRowHeight(row, tallest_row_height)

        self.total_purchase_label.setText(f"{total_purchase:.2f}")
        self.total_sales_label.setText(f"{total_sales:.2f}")
        self.total_income_label.setText(f"{(total_sales - total_purchase):.2f}")

    def search_product(self):
        search_text = self.search_history.text().lower()

        if not search_text:
            self.load_sales_for_today()
        else:
            filtered_sales = []
            for order_id, details in self.order_sales.items():
                for product_name, quantity in details["products"]:
                    if search_text in product_name.lower():
                        filtered_sales.append((order_id, details))
                        break
            self.update_sales_table(filtered_sales)

    def update_sales_table(self, filtered_sales):
        self.sales_table.setRowCount(0)
        tallest_row_height = 0#track yung tallest row

        for order_id, details in filtered_sales:
            products = "\n".join([p[0] for p in details["products"]])#\n for line  breaks
            quantities = "\n".join([str(p[1]) for p in details["products"]])
            total = details["total_sales"]
            sales_date = details["sales_date"]

            row_position = self.sales_table.rowCount()
            self.sales_table.insertRow(row_position)

            product_item = QTableWidgetItem(products)
            product_item.setTextAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft)
            self.sales_table.setItem(row_position, 1, product_item)

            self.sales_table.setItem(row_position, 0, QTableWidgetItem(str(order_id)))
            self.sales_table.setItem(row_position, 2, QTableWidgetItem(quantities))
            self.sales_table.setItem(row_position, 3, QTableWidgetItem(f"{total:.2f}"))
            self.sales_table.setItem(row_position, 4, QTableWidgetItem(str(sales_date)))

            self.sales_table.resizeRowsToContents()
            row_height = self.sales_table.rowHeight(row_position)
            if row_height > tallest_row_height:
                tallest_row_height = row_height

        #kada rows same sa tallest  row ang  allignment
        for row in range(self.sales_table.rowCount()):
            self.sales_table.setRowHeight(row, tallest_row_height)
    #excel printing
    def export_to_excel(self):
        if self.sales_table.rowCount() == 0:
            QMessageBox.warning(self, "No Data", "No sales data to export.")
            return

        selected_date = self.calendar.selectedDate().toString("yyyy-MM-dd")
        filename = f"sales_history_{selected_date}.xlsx"
        path = os.path.join(os.path.expanduser("~"), filename)

        try:
            data = []
            for row in range(self.sales_table.rowCount()):
                row_data = []
                for column in range(self.sales_table.columnCount()):
                    item = self.sales_table.item(row, column)
                    text = item.text() if item else ""
                    #newlines for 30 exceeded char
                    if column == 1:
                        text = '\n'.join([text[i:i+30] for i in range(0, len(text), 30)])
                    row_data.append(text)
                data.append(row_data)

            df = pd.DataFrame(data, columns=[
                "Order ID", "Product Name", "Quantity", "Total Retail Sales", "Sales Date"
            ])
            df.to_excel(path, index=False)

            #textwrapping
            wb = openpyxl.load_workbook(path)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, max_col=5, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

            wb.save(path)
            QMessageBox.information(self, "Export Successful", f"Saved to: {path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))
    #pdf printing 
    def export_to_pdf(self):
        if not hasattr(self, "order_sales") or not self.order_sales:
            QMessageBox.warning(self, "No Data", "No sales data to export.")
            return

        selected_date = self.calendar.selectedDate().toString("yyyy-MM-dd")
        filename = f"sales_history_{selected_date}.pdf"
        path = os.path.join(os.path.expanduser("~"), filename)

        try:
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=10)

            line_height = 6
            col_widths = [25, 60, 30, 40, 35]
            headers = ["Order ID", "Product Name", "Quantity", "Total Retail Sales", "Sales Date"]

            for i, header in enumerate(headers):
                pdf.cell(col_widths[i], line_height, header, border=1)
            pdf.ln(line_height)

            for order_id, details in self.order_sales.items():
                products = details["products"]
                total_sales = details["total_sales"]
                sales_date = str(details["sales_date"])

                for i, (product, quantity) in enumerate(products):
                    row_data = [
                        str(order_id) if i == 0 else "",
                        product,
                        str(quantity),
                        f"{total_sales:.2f}" if i == 0 else "",
                        sales_date if i == 0 else ""
                    ]
                    for j, data in enumerate(row_data):
                        pdf.cell(col_widths[j], line_height, data, border=1)
                    pdf.ln(line_height)

            pdf.output(path)
            QMessageBox.information(self, "Export Successful", f"PDF saved to: {path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))


class SalesLoaderThread(QThread):
    finished = pyqtSignal(list)
    def __init__(self, user_id, selected_date, db_config):
        super().__init__()
        self.user_id = user_id
        self.selected_date = selected_date
        self.db_config = db_config
    #then select dito yung data sa db
    def run(self):
        try:
            conn = mariadb.connect(**self.db_config)
            cursor = conn.cursor()
            cursor.execute("""
                SELECT o.orderId, p.productName, od.quantity, od.totalPrice, o.orderDateTime, p.purchasePrice
                FROM order_details od
                JOIN orders o ON od.orderId = o.orderId
                JOIN products p ON od.productId = p.productId
                WHERE o.userId = ? AND DATE(o.orderDateTime) = ?
            """, (self.user_id, self.selected_date))
            sales_data = cursor.fetchall()
        except Exception as e:
            print("Error loading sales:", e)
            sales_data = []
        finally:
            if conn:
                cursor.close()
                conn.close()
            self.finished.emit(sales_data)

